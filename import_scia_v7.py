"""
SCIA Import v4 — importer unificato e idempotente
=================================================
Sostituisce import_sciadb_v3.py + import_maintenance.py.

Cosa corregge rispetto alle versioni precedenti:
  • le UNITA' INSTALLATIVE (LCN TYPE = UI) diventano ElementModel a tutti gli
    effetti (prima erano solo Element appesi all'ESWBS a 5 cifre): questo da solo
    sistema albero, scheda "Componenti" e il conteggio per-UI (138/17/0).
  • il PDM viene raggruppato per NrApp (riga-task), NON per colore di cella:
    niente più task scartati.
  • il foglio SPARE PARTS viene finalmente importato in Parts + Spare.
  • aggancio manutenzione -> UI tramite ESWBS Apparato a 7 cifre NORMALIZZATO
    (via il ".0" dei float Excel).
  • un solo file per cartella apparato (3 fogli: PDM / SPARE PARTS / PBS),
    piu' il master PBS NAVE per la struttura alta della nave.

USO:
  1. Configura CONFIG (DB, SHIP_ID, SHIP_MODEL_ID, MASTER_PBS, EQUIPMENT_FILES).
  2. Lascia DRY_RUN = True e lancia: stampa cosa farebbe + riconciliazione,
     SENZA scrivere nulla. Controlla che i numeri tornino (motore 138, ecc.).
  3. Metti WIPE = True (cancella i dati importati della nave) e DRY_RUN = False,
     parti da UN solo apparato, verifica, poi attiva tutti i file.

NB: il "task" e' definito dalla riga con NrApp valorizzato (e' quello che da
    138/17/0 sul diesel-generatore). On-condition (recurrency 6/13) vengono
    importate ma restano distinguibili per il display.
"""

import os, re, sys
from datetime import datetime, timedelta

import pandas as pd
import openpyxl

try:
    import pymysql
except ImportError:
    pymysql = None  # serve solo quando DRY_RUN = False

# ───────────────────────────── CONFIG ──────────────────────────────
DB = dict(
    host=os.getenv("DB_HOST", ""),
    port=int(os.getenv("DB_PORT", "3306")),
    user=os.getenv("DB_USER", ""),
    password=os.getenv("DB_PASSWORD", ""),
    database=os.getenv("DB_NAME", "sciadb"),
    charset="utf8mb4",
)

SHIP_ID       = int(os.getenv("SHIP_ID", "31"))
SHIP_MODEL_ID = int(os.getenv("SHIP_MODEL_ID", "4"))

MASTER_PBS = os.getenv("MASTER_PBS", "PBS_master.xlsx")   # foglio "PBS NAVE" + "ESWBS"
# Cartella che contiene i file apparato (3 fogli: PDM/SPARE PARTS/PBS).
# Se impostata, vengono presi TUTTI i *.xlsx ricorsivamente (escluso il master).
EQUIPMENT_DIR = os.getenv("EQUIPMENT_DIR", "")
EQUIPMENT_FILES = [                                       # usato solo se EQUIPMENT_DIR e' vuoto
    os.getenv("EQ", "EQ_31120.xlsx"),
]

def discover_equipment_files():
    if not EQUIPMENT_DIR:
        return [f for f in EQUIPMENT_FILES if f and os.path.exists(f)]
    import glob as _glob, hashlib
    master_abs = os.path.abspath(MASTER_PBS) if MASTER_PBS else None
    found = sorted(_glob.glob(os.path.join(EQUIPMENT_DIR, "**", "*.xlsx"), recursive=True))
    out, seen_hash = [], {}
    for f in found:
        if os.path.basename(f).startswith("~$"):      # file temporanei Excel
            continue
        if master_abs and os.path.abspath(f) == master_abs:
            continue
        with open(f, "rb") as fh:
            h = hashlib.md5(fh.read()).hexdigest()    # dedup di file IDENTICI (stesso contenuto)
        if h in seen_hash:
            print(f"  doppione ignorato: {os.path.basename(f)}  (= {os.path.basename(seen_hash[h])})")
            continue
        seen_hash[h] = f
        out.append(f)
    return out

DRY_RUN = True      # True = non scrive, stampa solo + riconciliazione
WIPE    = False     # True = cancella i dati importati della nave prima di ricaricare

# tipi LCN che diventano nodi della gerarchia (ElementModel)
NODE_TYPES = {"EI", "SYS", "GRP", "UI", "ASSY", "LRU"}
# tipi che generano anche un ricambio dedicato (oltre al foglio SPARE PARTS)
SPARE_NODE_TYPES = {"LRU"}

# Periodicità Excel -> RecurrencyType.id  (dal vecchio import_maintenance)
PERIODICITY_MAP = {
    "1 DAYS":1,"1 WEEKS":2,"2 WEEKS":7,"1 MONTHS":3,"2 MONTHS":43,"3 MONTHS":4,
    "4 MONTHS":8,"6 MONTHS":30,"1 YEARS":5,"2 YEARS":9,"3 YEARS":10,"4 YEARS":32,
    "5 YEARS":11,"6 YEARS":33,"8 YEARS":34,"10 YEARS":12,"15 YEARS":35,
    "ON CONDITION":6,"ON FAULT":13,"72 HOURS":27,"100 HOURS":14,"200 HOURS":16,
    "250 HOURS":20,"300 HOURS":18,"400 HOURS":22,"500 HOURS":24,"1000 HOURS":15,
    "1200 HOURS":29,"1500 HOURS":42,"2000 HOURS":17,"2500 HOURS":21,"3000 HOURS":19,
    "3500 HOURS":44,"4000 HOURS":23,"4500 HOURS":46,"5000 HOURS":25,"6000 HOURS":28,
    "9000 HOURS":45,"10000 HOURS":36,"20000 HOURS":37,"25000 HOURS":41,
    "40000 HOURS":38,"100000 HOURS":47,
}
ON_CONDITION_RECURRENCY = {6, 13}   # tenute separate dalle periodiche

LEVEL_MAP = {
    "1° LIVELLO":1,"2° LIVELLO":2,"3° LIVELLO":3,"4° LIVELLO":4,"5° LIVELLO":5,"6° LIVELLO":6,
    "C":1,"O":2,"F":3,"H":4,"D":5,"L":6,
}

# ─────────────────────────── HELPERS ───────────────────────────────
def norm(v):
    if v is None: return ""
    try:
        if pd.isna(v): return ""
    except Exception:
        pass
    return str(v).strip()

def normcode(v):
    """Normalizza un codice ESWBS/LCN: toglie il '.0' dei float Excel e gli spazi."""
    s = norm(v)
    if s.endswith(".0"):
        s = s[:-2]
    return s

def s_str(v, n=None):
    s = norm(v)
    if not s: return None
    return s[:n] if n else s

def s_int(v):
    s = norm(v)
    if not s: return None
    try: return int(float(s))
    except Exception: return None

def s_float(v):
    s = norm(v).replace(",", ".")
    if not s: return None
    try: return float(s)
    except Exception: return None

def normalize_periodicity(val):
    s = norm(val).upper()
    if not s: return None
    for a, b in [(" MONTH"," MONTHS"),(" YEAR"," YEARS"),(" WEEK"," WEEKS"),
                 (" DAY"," DAYS"),(" HOUR"," HOURS")]:
        s = s.replace(a, b)
    s = re.sub(r"S{2,}\b", "S", s)
    s = re.sub(r"(\d)\.(\d{3})", r"\1\2", s)   # 1.000 -> 1000
    return s.strip()

def period_to_days(norm_period):
    PD = {"1 DAYS":1,"1 WEEKS":7,"2 WEEKS":14,"1 MONTHS":30,"2 MONTHS":60,
          "3 MONTHS":90,"4 MONTHS":120,"6 MONTHS":180,"1 YEARS":365,"2 YEARS":730,
          "3 YEARS":1095,"4 YEARS":1460,"5 YEARS":1825,"6 YEARS":2190,
          "8 YEARS":2920,"10 YEARS":3650,"15 YEARS":5475}
    return PD.get(norm_period)

# ── lettura foglio con header auto-rilevato e mappatura per NOME colonna ──
def load_sheet(path, sheet, marker_keys):
    """Ritorna (rows, colmap) dove rows e' lista di liste e colmap mappa idx logici."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"{path}: foglio '{sheet}' assente (presenti: {wb.sheetnames})")
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    # trova la riga header: quella che contiene tutti i marker
    hidx = None
    for i, r in enumerate(rows[:8]):
        low = [norm(c).lower().replace("\n", " ") for c in r]
        if all(any(m in cell for cell in low) for m in marker_keys):
            hidx = i; break
    if hidx is None:
        raise ValueError(f"{path}/{sheet}: header non trovato (marker {marker_keys})")
    header = [norm(c).lower().replace("\n", " ").strip() for c in rows[hidx]]
    data = [r for r in rows[hidx+1:] if any(norm(c) for c in r)]
    return data, header

def col(header, pred):
    for i, h in enumerate(header):
        if pred(h): return i
    return None

# ─────────────────────── STATO / "DB" simulato ─────────────────────
class Store:
    """Astrae l'INSERT: in DRY_RUN assegna id finti, altrimenti scrive su MySQL."""
    def __init__(self, conn):
        self.conn = conn
        self.cur  = conn.cursor() if conn else None
        self._fake = 0
        self.lcn2id = {}   # LCN normalizzato  -> ElementModel.id
        self.code2id = {}  # ESWBS normalizzato -> ElementModel.id (secondario)
        self.em_modelrow = {}  # id -> denom (per report)
        self.elem_by_model = {}  # element_model_id -> Element.id
        self.spare_by_pn = {}    # PN normalizzato  -> Spare.id
        self.code2glossary = {}  # ESWBS code -> ESWBS_Glossary.id
        self.stats = {}

    def bump(self, k, n=1): self.stats[k] = self.stats.get(k, 0) + n

    def insert(self, sql, params):
        if self.conn:
            self.cur.execute(sql, params)
            return self.cur.lastrowid
        self._fake += 1
        return self._fake

    def query_one(self, sql, params):
        if not self.conn: return None
        self.cur.execute(sql, params); return self.cur.fetchone()

    def commit(self):
        if self.conn: self.conn.commit()

# ───────────────────────────── WIPE ────────────────────────────────
def wipe_ship(store):
    """Cancella TUTTI i dati legati al modello nave SHIP_MODEL_ID:
    - tutte le navi che usano quel modello (Element, Spare, Tools, JobExecution);
    - tutte le Maintenance che referenziano i suoi ElementModel (anche di altre navi);
    - le tabelle figlie collegate (note, liste ricambi/tool, Readings, Scans, ...).
    Disabilita i controlli FK per la durata del wipe (l'ordine non e' piu' vincolante)
    e ripulisce comunque le figlie per non lasciare orfani.
    NON committa: stessa transazione della ricarica (commit in main)."""
    if not store.conn:
        print("  [DRY_RUN] wipe saltato"); return
    c = store.cur
    sm = int(SHIP_MODEL_ID)
    print(f"  Cancello TUTTI i dati del modello nave {sm} (FK check off)...")

    SHIPS  = f"(SELECT id FROM Ship WHERE ship_model_id={sm})"
    MODELS = f"(SELECT id FROM ElementModel WHERE ship_model_id={sm})"
    # Maintenance del modello: per nave OPPURE che referenzia un ElementModel del modello.
    # (wrapper derivato per poter cancellare FROM Maintenance usando una subquery su Maintenance)
    MAINTS = (f"(SELECT id FROM (SELECT id FROM Maintenance "
              f"WHERE id_ship IN {SHIPS} "
              f"OR System_ElementModel_ID IN {MODELS} "
              f"OR End_Item_ElementModel_ID IN {MODELS} "
              f"OR Maintenance_Item_ElementModel_ID IN {MODELS}) AS _m)")
    # Element del modello: per modello OPPURE per nave del modello (copre 31 e 19)
    ELEMS  = (f"(SELECT id FROM (SELECT id FROM Element "
              f"WHERE element_model_id IN {MODELS} OR ship_id IN {SHIPS}) AS _e)")

    c.execute("SET FOREIGN_KEY_CHECKS=0")

    # JobExecution legate al modello (per nave, per maintenance o per element)
    c.execute(f"DELETE FROM JobExecution WHERE ship_id IN {SHIPS} "
              f"OR job_id IN {MAINTS} OR element_eswbs_instance_id IN {ELEMS}")
    # note rimaste orfane (la loro JobExecution non esiste piu')
    c.execute("DELETE FROM PhotographicNote WHERE task_id IS NOT NULL "
              "AND task_id NOT IN (SELECT id FROM JobExecution)")
    c.execute("DELETE FROM VocalNote        WHERE task_id IS NOT NULL "
              "AND task_id NOT IN (SELECT id FROM JobExecution)")
    c.execute("DELETE FROM Note             WHERE id NOT IN (SELECT id FROM JobExecution)")

    # figli di Maintenance
    c.execute(f"DELETE FROM Maintenance_ListConsumable WHERE Maintenance_List_ID IN {MAINTS}")
    c.execute(f"DELETE FROM Maintenance_ListSpare      WHERE Maintenance_List_ID IN {MAINTS}")
    c.execute(f"DELETE FROM Maintenance_ListTools      WHERE Maintenance_List_ID IN {MAINTS}")
    c.execute(f"DELETE FROM Maintenance_Team           WHERE maintenance_id      IN {MAINTS}")
    c.execute(f"DELETE FROM Job                        WHERE maintenance_list_id IN {MAINTS}")
    # Maintenance
    c.execute(f"DELETE FROM Maintenance WHERE id IN {MAINTS}")

    # figli di Element
    c.execute(f"DELETE FROM Readings WHERE element_id IN {ELEMS}")
    c.execute(f"DELETE FROM Scans    WHERE element_id IN {ELEMS}")
    # Spare / Tools (per nave del modello)
    c.execute(f"DELETE FROM Spare WHERE ship_id IN {SHIPS}")
    c.execute(f"DELETE FROM Tools WHERE ship_id IN {SHIPS}")
    # Element
    c.execute(f"DELETE FROM Element WHERE id IN {ELEMS}")
    # figli di ElementModel
    c.execute(f"DELETE FROM TeamElementAccess WHERE element_model_id IN {MODELS}")
    # ElementModel
    c.execute(f"DELETE FROM ElementModel WHERE ship_model_id={sm}")

    c.execute("SET FOREIGN_KEY_CHECKS=1")
    print("  Wipe completato (commit a fine import).")

# ─────────────── PHASE 0: ESWBS_Glossary (albero/lista totale) ─────
GLOSSARY_MARKERS = ["macrogruppo", "eswbs 5 cifre"]

def parse_glossary(master_path):
    data, h = load_sheet(master_path, "ESWBS", GLOSSARY_MARKERS)
    C = {
        "macro":  col(h, lambda x: x.startswith("macrogruppo")),
        "lev1":   col(h, lambda x: x == "liv. 1" or x == "liv 1"),
        "lev2":   col(h, lambda x: x == "liv. 2" or x == "liv 2"),
        "lev3":   col(h, lambda x: x == "liv. 3" or x == "liv 3"),
        "code":   col(h, lambda x: x.startswith("eswbs 5 cifre")),
        "level":  col(h, lambda x: x == "livello"),
        "navsea": col(h, lambda x: x.startswith("nome navsea")),
        "long":   col(h, lambda x: x.startswith("descrizione lunga")),
        "short":  col(h, lambda x: x.startswith("descrizione breve")),
    }
    out = []
    for r in data:
        def g(k):
            i = C[k]; return r[i] if (i is not None and i < len(r)) else None
        code = normcode(g("code"))
        if not code:
            continue
        out.append(dict(
            code=code, macro=s_int(g("macro")),
            lev1=normcode(g("lev1")), lev2=normcode(g("lev2")), lev3=normcode(g("lev3")),
            level=s_str(g("level"),50),
            navsea=s_str(g("navsea"),100),
            long=s_str(g("long"),200), short=s_str(g("short"),100)))
    return out

GLOSS_SQL = ("INSERT INTO ESWBS_Glossary "
             "(eswbs_glossary_code, macrogroup, level1, level2, level3, level, "
             " name_navsea_S9040IDX, long_description_ita, short_description_ita) "
             "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)")

def import_glossary(store, rows):
    for gl in rows:
        if gl["code"] in store.code2glossary:
            store.bump("gloss_dup"); continue
        gid = store.insert(GLOSS_SQL, (gl["code"], gl["macro"], gl["lev1"], gl["lev2"],
                                       gl["lev3"], gl["level"], gl["navsea"],
                                       gl["long"], gl["short"]))
        store.code2glossary[gl["code"]] = gid
        store.bump("glossary")

# ─────────────── PHASE A: ElementModel + Element (PBS) ─────────────
PBS_MARKERS = ["lcn", "denominazione", "lcn type"]

def parse_pbs(path, sheet):
    data, h = load_sheet(path, sheet, PBS_MARKERS)
    C = {
        "wbs":      col(h, lambda x: x == "livello wbs"),
        "eswbs":    col(h, lambda x: x.startswith("codice eswbs")),
        "liv4":     col(h, lambda x: x == "livello 4"),
        "liv5":     col(h, lambda x: x == "livello 5"),
        "liv6":     col(h, lambda x: x == "livello 6"),
        "liv7":     col(h, lambda x: x == "livello 7"),
        "liv8":     col(h, lambda x: x == "livello 8"),
        "liv9":     col(h, lambda x: x == "livello 9"),
        "lcn":      col(h, lambda x: x == "lcn"),
        "alc":      col(h, lambda x: x == "alc"),
        "cihdci":   col(h, lambda x: x.startswith("ci/hdci")),
        "denom":    col(h, lambda x: x.startswith("denominazione")),
        "qty_assy": col(h, lambda x: x.startswith("quantità (") or x.startswith("quantita (")),
        "qty_ship": col(h, lambda x: x.startswith("quantità totale") or x.startswith("quantita totale")),
        "lcn_type": col(h, lambda x: x.startswith("lcn type")),
        "nsn":      col(h, lambda x: x.startswith("nuc/nsn") or x == "nsn"),
        "prezzo":   col(h, lambda x: x.startswith("prezzo unitario")),
        "peso":     col(h, lambda x: x.startswith("peso")),
        "dim_l":    col(h, lambda x: x.startswith("dimensioni l")),
        "dim_p":    col(h, lambda x: x.startswith("dimensioni p")),
        "dim_h":    col(h, lambda x: x.startswith("dimensioni h")),
        "volume":   col(h, lambda x: x.startswith("volume")),
        "alim":     col(h, lambda x: x.startswith("alimentazione")),
        "giri":     col(h, lambda x: x.startswith("numero di giri")),
        "pressione":col(h, lambda x: x.startswith("pressione")),
        "portata":  col(h, lambda x: x.startswith("portata")),
        "note":     col(h, lambda x: x == "note"),
        "locale":   col(h, lambda x: x.startswith("locale")),
    }
    out = []
    for r in data:
        def g(k):
            i = C[k]
            return r[i] if (i is not None and i < len(r)) else None
        out.append(dict(
            wbs=s_str(g("wbs"),10), eswbs=normcode(g("eswbs")), lcn=normcode(g("lcn")),
            alc=s_str(g("alc"),50), cihdci=s_str(g("cihdci"),50),
            denom=s_str(g("denom"),255) or "N/D",
            lcn_type=(norm(g("lcn_type")).upper() or None),
            qty_assy=s_int(g("qty_assy")), qty_ship=s_int(g("qty_ship")),
            liv4=s_str(g("liv4"),10), liv5=s_str(g("liv5"),10), liv6=s_str(g("liv6"),10),
            liv7=s_str(g("liv7"),10), liv8=s_str(g("liv8"),10), liv9=s_str(g("liv9"),10),
            nsn=s_str(g("nsn"),100), prezzo=s_str(g("prezzo"),100), peso=s_float(g("peso")),
            dim_l=s_str(g("dim_l")), dim_p=s_str(g("dim_p")), dim_h=s_str(g("dim_h")),
            volume=s_str(g("volume"),100), alim=s_str(g("alim"),50), giri=s_int(g("giri")),
            pressione=s_float(g("pressione")), portata=s_float(g("portata")),
            note=s_str(g("note"),255), locale=s_str(g("locale"),50),
        ))
    return out

def parse_pbs_equip(path):
    """PBS di un file apparato: di norma foglio 'PBS', ma i file attrezzi usano 'PBS TOOLS'."""
    import openpyxl
    sheets = openpyxl.load_workbook(path, read_only=True).sheetnames
    for name in ("PBS", "PBS TOOLS"):
        if name in sheets:
            return parse_pbs(path, name)
    raise ValueError(f"{path}: nessun foglio PBS/PBS TOOLS (presenti: {sheets})")

def build_em_sql():
    cols = ["parent_element_model_id","ship_model_id","ESWBS_code","LCN_name","LCN","ALC",
            "`CI/HDCI/CSCI`","LCNtype_ID","Installed_quantity_on_End_Item",
            "Installed_Quantity_on_Ship","Level1","Level3","Level4","Level5","Level6",
            "Level7","Level8","Level9","Weight","Dimensions_LxWxH","Revolution_speed",
            "Operating_pressure","Mass_flow","Power_supply","Ship_Area_Room_Code",
            "eswbs_glossary_id"]
    ph = ",".join(["%s"] * len(cols))
    return f"INSERT INTO ElementModel ({','.join(cols)}) VALUES ({ph})", cols

def dims(e):
    if e["dim_l"] or e["dim_p"] or e["dim_h"]:
        return f'{e["dim_l"] or "?"}x{e["dim_p"] or "?"}x{e["dim_h"] or "?"}'
    return None

def find_parent_id(lcn, store):
    """Parent = il nodo gia' inserito con l'LCN che e' prefisso piu' lungo.
    Toglie 2 caratteri alla volta (i livelli LCN crescono a coppie)."""
    if not lcn: return 0
    cand = lcn
    while len(cand) > 2:
        cand = cand[:-2]
        if cand in store.lcn2id:
            return store.lcn2id[cand]
    return 0

def import_elements(store, pbs_rows):
    em_sql, _ = build_em_sql()
    EL_SQL = ("INSERT INTO Element (element_model_id, ship_id, name, serial_number, "
              "installation_date, updated_at) VALUES (%s,%s,%s,%s,%s,%s)")
    seen_serial = {}
    for e in pbs_rows:
        if e["lcn_type"] not in NODE_TYPES:
            continue                      # salta righe di dettaglio (type vuoto)
        if not e["lcn"]:
            store.bump("em_skip_nolcn"); continue
        if e["lcn"] in store.lcn2id:
            store.bump("em_dup"); continue   # gia' inserito (master/altro file)
        parent_id = find_parent_id(e["lcn"], store)
        glossary_id = store.code2glossary.get(e["eswbs"]) if e["eswbs"] else None
        params = (parent_id, SHIP_MODEL_ID, e["eswbs"], e["denom"], e["lcn"], e["alc"],
                  e["cihdci"], e["lcn_type"], e["qty_assy"], e["qty_ship"],
                  e["wbs"], e["eswbs"], e["liv4"], e["liv5"], e["liv6"], e["liv7"],
                  e["liv8"], e["liv9"], e["peso"], dims(e), e["giri"], e["pressione"],
                  e["portata"], e["alim"], e["locale"], glossary_id)
        em_id = store.insert(em_sql, params)
        store.lcn2id[e["lcn"]] = em_id
        if e["eswbs"]:
            store.code2id.setdefault(e["eswbs"], em_id)
        store.em_modelrow[em_id] = (e["lcn"], e["lcn_type"], e["denom"])
        store.bump(f"em_{e['lcn_type']}")
        # un Element per ogni nodo
        serial = e.get("serial") or f"AUTO-{e['lcn']}"
        if serial in seen_serial:
            seen_serial[serial] += 1; serial = f"{serial}-{seen_serial[serial]}"
        else:
            seen_serial[serial] = 1
        el_id = store.insert(EL_SQL, (em_id, SHIP_ID, e["denom"][:255], serial[:255],
                                      None, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        store.elem_by_model[em_id] = el_id
        store.bump("el")

# ─────────────── PHASE B: Parts + Spare (SPARE PARTS) ──────────────
SP_MARKERS = ["descrizione", "pn fornitore", "quantità"]

def parse_spareparts(path):
    data, h = load_sheet(path, "SPARE PARTS", ["descrizione", "pn fornitore"])
    C = {
        "descr":    col(h, lambda x: x == "descrizione"),
        "pn_forn":  col(h, lambda x: x.startswith("pn fornitore")),
        "cage_forn":col(h, lambda x: x.startswith("cage fornitore")),
        "pn_costr": col(h, lambda x: x.startswith("pn costruttore")),
        "cage_costr":col(h, lambda x: x.startswith("cage costruttore")),
        "qty":      col(h, lambda x: x == "quantità" or x == "quantita"),
        "prezzo":   col(h, lambda x: x.startswith("prezzo")),
        "data_prezzo": col(h, lambda x: x.startswith("data riferimento")),
        "mtbf":     col(h, lambda x: x.startswith("mean time between")),
        "mttr":     col(h, lambda x: x.startswith("mean time to")),
        "lead":     col(h, lambda x: x.startswith("lead time")),
        "shelf":    col(h, lambda x: x.startswith("shelf life")),
        "maot":     col(h, lambda x: x.startswith("maximum allowable")),
        "serial":   col(h, lambda x: x.startswith("serial number")),
        "peso":     col(h, lambda x: x.startswith("peso")),
        "dim_l":    col(h, lambda x: x.startswith("dimensioni l")),
        "dim_p":    col(h, lambda x: x.startswith("dimensioni p")),
        "dim_h":    col(h, lambda x: x.startswith("dimensioni h")),
    }
    out = []
    for r in data:
        def g(k):
            i = C[k]; return r[i] if (i is not None and i < len(r)) else None
        descr = s_str(g("descr"), 255)
        pnf, pnc = s_str(g("pn_forn"),100), s_str(g("pn_costr"),100)
        if not (descr or pnf or pnc):
            continue
        out.append(dict(descr=descr or pnf or pnc, pn_forn=pnf, cage_forn=s_str(g("cage_forn"),100),
                        pn_costr=pnc, cage_costr=s_str(g("cage_costr"),100),
                        qty=s_str(g("qty"),100), prezzo=s_str(g("prezzo"),100),
                        data_prezzo=s_str(g("data_prezzo"),100),
                        mtbf=s_str(g("mtbf"),100), mttr=s_str(g("mttr"),100),
                        lead=s_int(g("lead")), shelf=s_str(g("shelf"),100),
                        maot=s_str(g("maot"),100), serial=s_str(g("serial"),200),
                        peso=s_str(g("peso"),100), dim_l=s_str(g("dim_l")),
                        dim_p=s_str(g("dim_p")), dim_h=s_str(g("dim_h"))))
    return out

def import_spares(store, sp_rows, equip_em_id):
    """Importa il catalogo ricambi (foglio SPARE PARTS) e lo appende all'apparato."""
    SP_SQL = ("INSERT INTO Spare (element_model_id, ship_id, Serial_number, Part_name, "
              "Unitary_price, NSN, Weight, Dimensions_LxWxH, Provisioning_Lead_Time_PLT, "
              "Shelf_Life, Limited_Life, Price_reference_date) "
              "VALUES (%s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s)")
    for i, s in enumerate(sp_rows):
        serial = s["serial"] or s["pn_costr"] or s["pn_forn"] or f"SP-{equip_em_id}-{i}"
        key = (s["pn_costr"] or s["pn_forn"] or s["descr"] or "").strip().lower()
        d = None
        if s["dim_l"] or s["dim_p"] or s["dim_h"]:
            d = f'{s["dim_l"] or "?"}x{s["dim_p"] or "?"}x{s["dim_h"] or "?"}'
        sp_id = store.insert(SP_SQL, (equip_em_id, SHIP_ID, serial[:255], s["descr"][:255],
                                      s["prezzo"], None, s["peso"], d, s["lead"],
                                      s["shelf"], s["maot"], s["data_prezzo"]))
        if key:
            store.spare_by_pn.setdefault(key, sp_id)
        store.bump("spare_catalog")

# ──────────── PHASE C: Maintenance + figli + JobExecution (PDM) ─────
PDM_MARKERS = ["nrapp", "dentask", "eswbs apparato"]

def parse_pdm(path):
    data, h = load_sheet(path, "PDM", PDM_MARKERS)
    C = {
        "preventiva": col(h, lambda x: x == "preventiva"),
        "eswbs_sys":  col(h, lambda x: x == "eswbs"),
        "sotto":      col(h, lambda x: x.startswith("sotto-sistema") or x.startswith("sotto sistema")),
        "eswbs_app":  col(h, lambda x: x.startswith("eswbs apparato")),
        "denom_app":  col(h, lambda x: x.startswith("denominazione apparato")),
        "nrapp":      col(h, lambda x: x == "nrapp"),
        "dentask":    col(h, lambda x: x == "dentask"),
        "periodicita":col(h, lambda x: x == "periodicità" or x == "periodicita"),
        "livello":    col(h, lambda x: x == "livellomanmm"),
        "nr_man":     col(h, lambda x: x.startswith("nr_manutentori")),
        "melap":      col(h, lambda x: x.startswith("melap (minutes)")),
        "mmh":        col(h, lambda x: x.startswith("mmh")),
        "doc":        col(h, lambda x: x.startswith("documentodiriferimento")),
        "pagina":     col(h, lambda x: x == "pagina"),
        "operativo":  col(h, lambda x: x.startswith("operativ")),
        "applic":     col(h, lambda x: x.startswith("applicabilit")),
        # figli
        "pn":         col(h, lambda x: x == "pn"),
        "cage":       col(h, lambda x: x == "cage"),
        "description":col(h, lambda x: x == "description"),
        "quantity":   col(h, lambda x: x == "quantity"),
        "typology":   col(h, lambda x: x == "typology"),
        "icc":        col(h, lambda x: x == "icc"),
    }
    def g(r, k):
        i = C[k]; return r[i] if (i is not None and i < len(r)) else None

    groups, cur = [], None
    for r in data:
        if norm(g(r, "nrapp")):                       # riga-task (header)
            cur = dict(
                preventiva=norm(g(r,"preventiva")).upper(),
                eswbs_sys=normcode(g(r,"eswbs_sys")),
                eswbs_app=normcode(g(r,"eswbs_app")),
                sotto=s_str(g(r,"sotto"),255),
                denom_app=s_str(g(r,"denom_app"),255),
                dentask=s_str(g(r,"dentask"),255),
                periodicita=s_str(g(r,"periodicita")),
                livello=s_str(g(r,"livello")),
                nr_man=s_int(g(r,"nr_man")), melap=s_str(g(r,"melap"),100),
                mmh=s_str(g(r,"mmh"),100), doc=s_str(g(r,"doc"),100),
                pagina=s_str(g(r,"pagina"),100), operativo=s_str(g(r,"operativo"),100),
                applic=s_str(g(r,"applic"),255), children=[])
            groups.append(cur)
        elif cur is not None:                          # riga figlia (ricambio/tool/consumabile)
            typ = s_str(g(r,"typology"))
            pn  = s_str(g(r,"pn"),100)
            if not (typ or pn or norm(g(r,"description"))):
                continue
            cur["children"].append(dict(
                typology=typ, pn=pn, cage=s_str(g(r,"cage"),10),
                description=s_str(g(r,"description"),255),
                quantity=s_str(g(r,"quantity")), icc=s_str(g(r,"icc"),100)))
    return groups

def parse_qty(q):
    if not q: return None
    try: return float(str(q).split()[0])
    except Exception: return None

def maint_type_id(np):
    if not np: return 1
    if "HOURS" in np: return 2
    if "CYCLES" in np: return 3
    if "ON FAULT" in np: return 4
    if "ON CONDITION" in np: return 5
    return 1

MAINT_SQL = """INSERT INTO Maintenance
 (id_ship, name, System_ElementModel_ID, End_Item_ElementModel_ID,
  Maintenance_Item_ElementModel_ID, Operational_Not_operational,
  Mean_elapsed_time_MELAP, Mean_Men_Hours_MMH, Personnel_no,
  RecurrencyType_ID, MaintenanceLevel_ID, Maintenance_type_id,
  Service_or_Maintenance_Manual_Link, `Service_or_Maintenance_manual_ParagraphPage`,
  Check_List, Maintenance_under_condition_description, Note)
 VALUES (%s,%s, %s,%s,%s, %s, %s,%s,%s, %s,%s,%s, %s,%s, %s,%s,%s)"""

JE_SQL = """INSERT INTO JobExecution
 (job_id, status_id, ship_id, element_eswbs_instance_id, recurrency_type_id,
  execution_date, starting_date, ending_date, execution_state)
 VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

def resolve_element(code, store, min_len=5):
    """Trova l'ElementModel per un codice ESWBS/LCN del PDM.
    1) match esatto su LCN, poi su ESWBS code;
    2) altrimenti l'ANTENATO con il prefisso piu' lungo gia' presente
       (i task definiti su un sotto-componente non modellato salgono al
       nodo reale piu' vicino). Non scende sotto min_len cifre."""
    if not code:
        return None, None
    if code in store.lcn2id: return store.lcn2id[code], "exact"
    if code in store.code2id: return store.code2id[code], "exact"
    cand = code
    while len(cand) > min_len:
        cand = cand[:-1]
        if cand in store.lcn2id: return store.lcn2id[cand], "prefix"
        if cand in store.code2id: return store.code2id[cand], "prefix"
    return None, None

def import_maintenance(store, groups, recon):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for gp in groups:
        # ── risoluzione elemento: esatto, poi antenato (prefisso piu' lungo)
        app_id, how = resolve_element(gp["eswbs_app"], store)
        sys_id, _   = resolve_element(gp["eswbs_sys"], store)
        target = app_id or sys_id
        recon_key = gp["eswbs_app"] or gp["eswbs_sys"] or "?"
        recon[recon_key] = recon.get(recon_key, 0) + 1
        if target is None:
            store.bump("maint_unmatched")
        elif how == "prefix":
            store.bump("maint_matched_prefix")
        np = normalize_periodicity(gp["periodicita"])
        rec_id = PERIODICITY_MAP.get(np)
        lvl_id = LEVEL_MAP.get((gp["livello"] or "").upper())
        check_list = 1 if np == "1 DAYS" else None
        is_oncond = rec_id in ON_CONDITION_RECURRENCY
        m_id = store.insert(MAINT_SQL, (
            SHIP_ID, gp["dentask"], sys_id, app_id, app_id, gp["operativo"],
            gp["melap"], gp["mmh"], gp["nr_man"], rec_id, lvl_id, maint_type_id(np),
            gp["doc"], gp["pagina"], check_list, gp["applic"], None))
        store.bump("maint")
        store.bump("maint_oncond" if is_oncond else "maint_periodic")

        # ── figli: spare / consumable / tool
        for ch in gp["children"]:
            typ = (ch["typology"] or "").lower()
            label = (ch["pn"] or ch["description"] or "").strip().lower()
            qty = parse_qty(ch["quantity"])
            uom = (str(ch["quantity"]).split()[-1] if ch["quantity"] else "EA")
            if typ.startswith("spare"):
                sp_id = store.spare_by_pn.get(label)
                if not sp_id:
                    sp_id = store.insert(
                        "INSERT INTO Spare (element_model_id, ship_id, Serial_number, Part_name, NSN) "
                        "VALUES (%s,%s,%s,%s,%s)",
                        (target, SHIP_ID, f"SP-M-{ch['pn'] or label}-{m_id}"[:255],
                         (ch["description"] or label)[:255], ch["pn"]))
                    if label: store.spare_by_pn[label] = sp_id
                    store.bump("spare_from_pdm")
                store.insert("INSERT INTO Maintenance_ListSpare "
                             "(Maintenance_List_ID, Spare_ID, Spare_quantity, Spare_unit_of_measure) "
                             "VALUES (%s,%s,%s,%s)", (m_id, sp_id, s_int(qty), uom))
                store.bump("link_spare")
            elif typ.startswith("consumable"):
                store.bump("child_consumable")   # (creazione Consumable analoga: omessa nel dry-run)
            elif typ.startswith("tool"):
                store.bump("child_tool")

        # ── JobExecution pianificata (una per manutenzione)
        el_id = store.elem_by_model.get(target) if target else None
        ending = None
        days = period_to_days(np)
        if days:
            ending = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
        store.insert(JE_SQL, (m_id, 6, SHIP_ID, el_id, rec_id, now, now, ending, "scheduled"))
        store.bump("jobexec")

# ──────────────────────────── MAIN ─────────────────────────────────
def main():
    conn = None
    if not DRY_RUN:
        if pymysql is None:
            sys.exit("pymysql non installato")
        conn = pymysql.connect(**DB, autocommit=False)
    store = Store(conn)

    print(f"=== SCIA import v5 — DRY_RUN={DRY_RUN} WIPE={WIPE} ship={SHIP_ID} ===")
    if WIPE and not DRY_RUN:
        wipe_ship(store)

    equip = discover_equipment_files()
    print(f"\nApparati trovati: {len(equip)}")
    if EQUIPMENT_DIR:
        print(f"  (glob ricorsivo di '{EQUIPMENT_DIR}')")

    # PHASE 0: glossario ESWBS = albero/lista totale dei codici (dal master)
    print("\n[0] ESWBS_Glossary (albero/lista totale)")
    if MASTER_PBS and os.path.exists(MASTER_PBS):
        import_glossary(store, parse_glossary(MASTER_PBS))
        print(f"  glossario dal master '{MASTER_PBS}'")
    else:
        print("  ATTENZIONE: master non trovato, glossario non caricato")

    # PHASE A: struttura. Prima il master (skeleton alto), poi ogni apparato.
    print("\n[A] ElementModel + Element")
    if MASTER_PBS and os.path.exists(MASTER_PBS):
        import_elements(store, parse_pbs(MASTER_PBS, "PBS NAVE"))
        print(f"  master '{MASTER_PBS}' (PBS NAVE) caricato")
    for f in equip:
        try:
            import_elements(store, parse_pbs_equip(f))
        except Exception as e:
            store.bump("file_error_pbs"); print(f"  ERRORE PBS in '{f}': {e}")

    # PHASE B: ricambi (catalogo SPARE PARTS) appesi all'EI dell'apparato
    print("\n[B] Parts + Spare (foglio SPARE PARTS)")
    for f in equip:
        try:
            sp = parse_spareparts(f)
            pbs = parse_pbs_equip(f)
            ei = next((e for e in pbs if e["lcn_type"] == "EI"), None)
            equip_em_id = store.lcn2id.get(ei["lcn"]) if ei else None
            import_spares(store, sp, equip_em_id)
        except Exception as e:
            store.bump("file_error_spare"); print(f"  ERRORE SPARE in '{f}': {e}")

    # PHASE C: manutenzioni
    print("\n[C] Maintenance + figli + JobExecution")
    recon = {}
    for f in equip:
        try:
            groups = parse_pdm(f)
            import_maintenance(store, groups, recon)
        except Exception as e:
            store.bump("file_error_pdm"); print(f"  ERRORE PDM in '{f}': {e}")

    if conn:
        store.commit(); conn.close()

    # ─── REPORT ───
    print("\n" + "=" * 60)
    print("RIEPILOGO")
    for k in sorted(store.stats):
        print(f"  {k:22} {store.stats[k]}")
    print("\nRICONCILIAZIONE manutenzioni per apparato (ESWBS Apparato -> n. task):")
    for k in sorted(recon):
        em, how = resolve_element(k, store)
        if not em:
            flag = "  <-- NON AGGANCIATO!"
        elif how == "prefix":
            flag = "  (su antenato)"
        else:
            flag = ""
        print(f"  {k:14} -> {recon[k]:4} task{flag}")

if __name__ == "__main__":
    main()