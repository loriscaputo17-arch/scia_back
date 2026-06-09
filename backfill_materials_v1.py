"""
SCIA — backfill materiali (consumabili / utensili) + recupero spare orfani — v2
================================================================================
Come v1 ma VELOCE: precarica in memoria (poche query bulk) manutenzioni,
link gia' esistenti e catalogo; fa il match in RAM; tocca il DB SOLO per
scrivere. In DRY_RUN non scrive: conta soltanto -> anteprima immediata.
Stampa avanzamento per file. Idempotente.

NON distruttivo. NON re-importa. Crea solo:
  (1) link mancanti task<->consumabile -> Maintenance_ListConsumable
  (2) link mancanti task<->utensile    -> Maintenance_ListTools
  (3) (opzionale) element_model_id sugli Spare orfani, dall'apparato del file.

USO
  1. Config DB (env) + EQUIPMENT_DIR (cartella) oppure EQUIPMENT_FILES.
  2. DRY_RUN=True  -> stampa conteggi, non scrive.
  3. Numeri ok? DRY_RUN=False -> scrive in transazione unica.
"""

import os, sys, glob
from collections import Counter, defaultdict
import openpyxl

try:
    import pymysql
except ImportError:
    pymysql = None

# ───────────────────────────── CONFIG ──────────────────────────────
DB = dict(
    host=os.getenv("DB_HOST", ""),
    port=int(os.getenv("DB_PORT", "3306")),
    user=os.getenv("DB_USER", ""),
    password=os.getenv("DB_PASSWORD", ""),
    database=os.getenv("DB_NAME", "sciadb"),
    charset="utf8mb4",
)
SHIP_ID       = int(os.getenv("SHIP_ID", "31"))
SHIP_MODEL_ID = int(os.getenv("SHIP_MODEL_ID", "4"))

# Per il PILOTA su un file:
EQUIPMENT_DIR   = "scia_files"                                   # vuoto -> usa EQUIPMENT_FILES
MASTER_PBS      = ""                                   # nome del master da escludere (se in cartella)
EQUIPMENT_FILES = ["scia_files/31120 - DIESEL GENERATORI - WARTSILA.xlsx"]
# Per TUTTI gli apparati: EQUIPMENT_DIR = "scia_files"  (EQUIPMENT_FILES ignorato)

DRY_RUN = True
RECOVER_ORPHAN_SPARES = True

# ─────────────────────────── HELPERS ───────────────────────────────
def norm(v):
    if v is None: return ""
    return str(v).strip()
def normcode(v):
    s = norm(v)
    return s[:-2] if s.endswith(".0") else s
def s_str(v, n=None):
    s = norm(v)
    if not s: return None
    return s[:n] if n else s
def parse_qty(q):
    if not q: return None
    try: return float(str(q).split()[0])
    except Exception: return None
def uom_of(q):
    s = norm(q); return s.split()[-1] if s else "EA"

def load_sheet(path, sheet, marker_keys):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close(); raise ValueError(f"foglio '{sheet}' assente (presenti: {wb.sheetnames})")
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    hidx = None
    for i, r in enumerate(rows[:8]):
        low = [norm(c).lower().replace("\n", " ") for c in r]
        if all(any(m in cell for cell in low) for m in marker_keys):
            hidx = i; break
    if hidx is None:
        raise ValueError(f"{path}/{sheet}: header non trovato (marker {marker_keys})")
    header = [norm(c).lower().replace("\n", " ").strip() for c in rows[hidx]]
    data = [r for r in rows[hidx+1:] if any(norm(c) for c in r)]
    return data, header
def col(header, pred):
    for i, h in enumerate(header):
        if pred(h): return i
    return None

PDM_MARKERS = ["nrapp", "dentask", "eswbs apparato"]

def parse_pdm(path):
    sheets = openpyxl.load_workbook(path, read_only=True).sheetnames
    pdm_sheets = [s for s in sheets if s.strip().lower() == "pdm" or s.strip().lower().startswith("pdm ")]
    if not pdm_sheets:
        raise ValueError(f"{path}: nessun foglio PDM (presenti: {sheets})")
    groups = []
    for sh in pdm_sheets:                       # ← tutto il corpo sta DENTRO il for
        data, h = load_sheet(path, sh, PDM_MARKERS)
        C = {"eswbs_sys": col(h, lambda x: x == "eswbs"),
             "eswbs_app": col(h, lambda x: x.startswith("eswbs apparato")),
             "nrapp":     col(h, lambda x: x == "nrapp"),
             "dentask":   col(h, lambda x: x == "dentask"),
             "pn":        col(h, lambda x: x == "pn"),
             "description":col(h, lambda x: x == "description"),
             "quantity":  col(h, lambda x: x == "quantity"),
             "typology":  col(h, lambda x: x == "typology"),
             "icc":       col(h, lambda x: x == "icc")}
        def g(r, k, C=C):
            i = C[k]; return r[i] if (i is not None and i < len(r)) else None
        cur = None
        for r in data:
            if norm(g(r, "nrapp")):
                cur = dict(eswbs_sys=normcode(g(r, "eswbs_sys")),
                           eswbs_app=normcode(g(r, "eswbs_app")),
                           dentask=s_str(g(r, "dentask"), 255), children=[])
                groups.append(cur)
            elif cur is not None:
                typ = s_str(g(r, "typology")); pn = s_str(g(r, "pn"), 100)
                desc = s_str(g(r, "description"), 255)
                if not (typ or pn or desc): continue
                cur["children"].append(dict(typology=typ, pn=pn, description=desc,
                                            quantity=s_str(g(r, "quantity")),
                                            icc=s_str(g(r, "icc"), 100)))
    return groups

def parse_spareparts(path):
    try:
        data, h = load_sheet(path, "SPARE PARTS", ["descrizione", "pn fornitore"])
    except Exception:
        return []
    C = {"descr": col(h, lambda x: x == "descrizione")}
    out = []
    for r in data:
        i = C["descr"]; descr = s_str(r[i], 255) if (i is not None and i < len(r)) else None
        if descr: out.append(descr)
    return out

def resolve_element(code, lcn2id, code2id, min_len=5):
    if not code: return None
    if code in lcn2id:  return lcn2id[code]
    if code in code2id: return code2id[code]
    cand = code
    while len(cand) > min_len:
        cand = cand[:-1]
        if cand in lcn2id:  return lcn2id[cand]
        if cand in code2id: return code2id[cand]
    return None

# ───────────────── PRECARICAMENTO IN MEMORIA (bulk) ────────────────
def preload(cur):
    print("  preload ElementModel...", flush=True)
    lcn2id, code2id = {}, {}
    cur.execute("SELECT id, LCN, ESWBS_code FROM ElementModel WHERE ship_model_id=%s", (SHIP_MODEL_ID,))
    for _id, lcn, code in cur.fetchall():
        if lcn:  lcn2id.setdefault(normcode(lcn), _id)
        if code: code2id.setdefault(normcode(code), _id)

    print("  preload Maintenance (indice name+End_Item)...", flush=True)
    maint_index = defaultdict(list)
    cur.execute("SELECT id, name, End_Item_ElementModel_ID FROM Maintenance WHERE id_ship=%s", (SHIP_ID,))
    for mid, name, eitem in cur.fetchall():
        maint_index[(name, eitem)].append(mid)

    print("  preload catalogo Consumables/Tools...", flush=True)
    cons_key, tool_key = {}, {}
    cur.execute("SELECT ID, ConsumableArticleCode, Commercial_Name FROM Consumables")
    for cid, codec, name in cur.fetchall():
        for k in (codec, name):
            k = (k or "").strip().lower()
            if k: cons_key.setdefault(k, cid)
    cur.execute("SELECT ID, Part_Number_OEM, Tool_name FROM Tools")
    for tid, pn, name in cur.fetchall():
        for k in (pn, name):
            k = (k or "").strip().lower()
            if k: tool_key.setdefault(k, tid)

    print("  preload link esistenti...", flush=True)
    cons_links = set()
    cur.execute("SELECT Maintenance_List_ID, Consumable_ID FROM Maintenance_ListConsumable")
    for m, c in cur.fetchall(): cons_links.add((m, c))
    tool_links = set()
    cur.execute("SELECT Maintenance_List_ID, `Tool ID` FROM Maintenance_ListTools")
    for m, t in cur.fetchall(): tool_links.add((m, t))

    orphan_by_name = defaultdict(list)
    if RECOVER_ORPHAN_SPARES:
        print("  preload Spare orfani...", flush=True)
        cur.execute("SELECT ID, Part_name FROM Spare WHERE ship_id=%s AND element_model_id IS NULL", (SHIP_ID,))
        for sid, pname in cur.fetchall():
            if pname: orphan_by_name[pname.strip().lower()].append(sid)

    return dict(lcn2id=lcn2id, code2id=code2id, maint_index=maint_index,
                cons_key=cons_key, tool_key=tool_key,
                cons_links=cons_links, tool_links=tool_links,
                orphan_by_name=orphan_by_name)

# ─────────────────────────── MAIN ──────────────────────────────────
def discover_files():
    if not EQUIPMENT_DIR:
        return [f for f in EQUIPMENT_FILES if f and os.path.exists(f)]
    import hashlib
    master_abs = os.path.abspath(MASTER_PBS) if MASTER_PBS else None
    found = sorted(glob.glob(os.path.join(EQUIPMENT_DIR, "**", "*.xlsx"), recursive=True))
    out, seen = [], {}
    for f in found:
        if os.path.basename(f).startswith("~$"): continue
        if master_abs and os.path.abspath(f) == master_abs: continue
        with open(f, "rb") as fh: h = hashlib.md5(fh.read()).hexdigest()
        if h in seen:
            print(f"  doppione ignorato: {os.path.basename(f)} (= {os.path.basename(seen[h])})"); continue
        seen[h] = f; out.append(f)
    return out

def main():
    files = discover_files()
    print(f"=== BACKFILL v2 — DRY_RUN={DRY_RUN} — file: {len(files)} ===", flush=True)
    if not files:
        print("Nessun file. Controlla EQUIPMENT_DIR / EQUIPMENT_FILES."); return
    if pymysql is None or not DB["host"]:
        print("Nessuna connessione DB (pymysql assente o DB_HOST vuoto)."); return

    conn = pymysql.connect(**DB, autocommit=False); cur = conn.cursor()
    print("Precaricamento dati dal DB...", flush=True)
    M = preload(cur)
    print(f"  -> ElementModel(LCN)={len(M['lcn2id'])} manutenzioni-indice={len(M['maint_index'])} "
          f"consumabili={len(M['cons_key'])} tool={len(M['tool_key'])} "
          f"orfani-nome={len(M['orphan_by_name'])}", flush=True)

    stats = Counter()
    fake = [0]
    def write(sql, params):
        if DRY_RUN:
            fake[0] -= 1; return fake[0]
        cur.execute(sql, params); return cur.lastrowid

    for idx, path in enumerate(files, 1):
        try:
            groups = parse_pdm(path)
        except Exception as e:
            print(f"[{idx}/{len(files)}] ERRORE PDM {os.path.basename(path)}: {e}"); continue
        print(f"[{idx}/{len(files)}] {os.path.basename(path)} — {len(groups)} task", flush=True)

        for gp in groups:
            app_id = resolve_element(gp["eswbs_app"], M["lcn2id"], M["code2id"])
            target = app_id or resolve_element(gp["eswbs_sys"], M["lcn2id"], M["code2id"])
            maint_ids = M["maint_index"].get((gp["dentask"], app_id), [])
            if not maint_ids and gp["children"]:
                stats["task_senza_maintenance"] += 1

            for ch in gp["children"]:
                typ = (ch["typology"] or "").lower()
                label = (ch["pn"] or ch["description"] or "").strip().lower()
                if not label: continue
                qty = parse_qty(ch["quantity"]); uom = uom_of(ch["quantity"])

                if typ.startswith("consumable"):
                    stats["src_consumable"] += 1
                    cid = M["cons_key"].get(label)
                    if not cid:
                        cid = write("INSERT INTO Consumables (Commercial_Name, ConsumableArticleCode, "
                                    "Consumable_quantity, ICC_Item_Category_Code) VALUES (%s,%s,%s,%s)",
                                    ((ch["description"] or label)[:100], ch["pn"],
                                     str(qty) if qty is not None else None, ch["icc"]))
                        M["cons_key"][label] = cid; stats["consumable_creati"] += 1
                    for m_id in maint_ids:
                        if (m_id, cid) in M["cons_links"]:
                            stats["link_consumable_gia_presente"] += 1; continue
                        write("INSERT INTO Maintenance_ListConsumable (Maintenance_List_ID, Consumable_ID, "
                              "Consumable_quantity, Consumable_quantity_Unit_of_measure) VALUES (%s,%s,%s,%s)",
                              (m_id, cid, qty, uom))
                        M["cons_links"].add((m_id, cid)); stats["link_consumable_creati"] += 1

                elif typ.startswith("tool"):
                    stats["src_tool"] += 1
                    tid = M["tool_key"].get(label)
                    if not tid:
                        tid = write("INSERT INTO Tools (element_model_id, ship_id, Tool_name, Part_Number_OEM, "
                                    "quantity, ICC_Item_Category_Code) VALUES (%s,%s,%s,%s,%s,%s)",
                                    (target, SHIP_ID, (ch["description"] or label)[:255], ch["pn"],
                                     str(qty) if qty is not None else None, ch["icc"]))
                        M["tool_key"][label] = tid; stats["tool_creati"] += 1
                    for m_id in maint_ids:
                        if (m_id, tid) in M["tool_links"]:
                            stats["link_tool_gia_presente"] += 1; continue
                        write("INSERT INTO Maintenance_ListTools (Maintenance_List_ID, `Tool ID`, "
                              "Tool_quantity, Tool_Quantity_Unit_of_measure) VALUES (%s,%s,%s,%s)",
                              (m_id, tid, str(qty) if qty is not None else None, uom))
                        M["tool_links"].add((m_id, tid)); stats["link_tool_creati"] += 1

        # (3) recupero spare orfani -> apparato dominante del file
        if RECOVER_ORPHAN_SPARES:
            codes = sorted({g["eswbs_app"] for g in groups if g["eswbs_app"]})
            app_ids = [resolve_element(c, M["lcn2id"], M["code2id"]) for c in codes]
            app_ids = [a for a in app_ids if a]
            if app_ids:
                dominant = Counter(app_ids).most_common(1)[0][0]
                for name in set(parse_spareparts(path)):
                    ids = M["orphan_by_name"].get(name.strip().lower(), [])
                    for sid in ids:
                        if not DRY_RUN:
                            cur.execute("UPDATE Spare SET element_model_id=%s "
                                        "WHERE ID=%s AND element_model_id IS NULL", (dominant, sid))
                        stats["spare_orfani_recuperati"] += 1
                    if ids:  # non riassegnare lo stesso orfano da un altro file
                        M["orphan_by_name"][name.strip().lower()] = []

    print("\n" + "=" * 56 + "\nRIEPILOGO", flush=True)
    for k in sorted(stats): print(f"  {k:34} {stats[k]}")

    if DRY_RUN:
        conn.rollback(); print("\nDRY_RUN: rollback, nessuna scrittura.")
    else:
        conn.commit(); print("\nCOMMIT eseguito.")
    conn.close()

if __name__ == "__main__":
    main()