"""
SCIADB Import Maintenance Script
Legge un file Excel PDM e importa in:
  - Maintenance
  - Consumables (se non esistono) + Maintenance_ListConsumable
  - Tools (se non esistono)       + Maintenance_ListTools
  - Spare (se non esistono)       + Maintenance_ListSpare

STRUTTURA EXCEL:
  - Riga verde (NrApp non-null) = manutenzione principale
  - Righe figlie (NrApp null)   = consumabili/tools/spare collegati

REGOLE:
  - Periodicità "1 DAYS" → Check_List = 1
  - TYPOLOGY = Consumable → Maintenance_ListConsumable
  - TYPOLOGY = Tool       → Maintenance_ListTools
  - TYPOLOGY = Spare      → Maintenance_ListSpare

ISTRUZIONI:
  1. Configura CONFIG
  2. python3 import_maintenance.py
  3. Puoi eseguire per più file Excel, lo script è idempotente sui Consumables/Tools
"""

import pandas as pd
import pymysql
import sys
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
DB_HOST="scia-project-questit.ccp0mjdczkug.eu-central-1.rds.amazonaws.com"
DB_NAME="sciadb"
DB_PASSWORD="MiTEwe64w6hAkIXiYo8aLavogIKO5i"
DB_PORT=3306
DB_USER="sciauser"

EXCEL_FILE  = "pdm64.xlsx"
SHIP_ID     = 31 

# Mappatura Periodicità Excel → RecurrencyType.id
PERIODICITY_MAP = {
    "1 DAYS":       1,   # Daily
    "1 WEEKS":      2,   # Weekly
    "2 WEEKS":      7,   # Every 2 weeks
    "1 MONTHS":     3,   # Monthly
    "2 MONTHS":    43,   # Every 2 months
    "3 MONTHS":     4,   # Quarterly
    "4 MONTHS":     8,   # Every 4 months
    "6 MONTHS":    30,   # Every 6 months
    "1 YEARS":      5,   # Yearly
    "2 YEARS":      9,   # Every 2 years
    "3 YEARS":     10,   # Every 3 years
    "4 YEARS":     32,   # Every 4 years
    "5 YEARS":     11,   # Every 5 years
    "6 YEARS":     33,   # Every 6 years
    "8 YEARS":     34,   # Every 8 years
    "10 YEARS":    12,   # Every 10 years
    "15 YEARS":    35,   # Every 15 years
    "ON CONDITION": 6,   # On condition
    "ON FAULT":    13,   # On fault - Troubleshooting
    "72 HOURS":    27,   # Every 72 hours
    "100 HOURS":   14,   # Every 100 hours
    "200 HOURS":   16,   # Every 200 hours
    "250 HOURS":   20,   # Every 250 hours
    "300 HOURS":   18,   # Every 300 hours
    "400 HOURS":   22,   # Every 400 hours
    "500 HOURS":   24,   # Every 500 hours
    "1000 HOURS":  15,   # Every 1.000 hours
    "1200 HOURS":  29,   # Every 1.200 hours
    "1500 HOURS":  42,   # Every 1.500 hours
    "2000 HOURS":  17,   # Every 2.000 hours
    "2500 HOURS":  21,   # Every 2.500 hours
    "3000 HOURS":  19,   # Every 3.000 hours
    "3500 HOURS":  44,   # Every 3.500 hours
    "4000 HOURS":  23,   # Every 4.000 hours
    "4500 HOURS":  46,   # Every 4.500 hours
    "5000 HOURS":  25,   # Every 5.000 hours
    "6000 HOURS":  28,   # Every 6.000 hours
    "9000 HOURS":  45,   # Every 9.000 hours
    "10000 HOURS": 36,   # Every 10.000 hours
    "20000 HOURS": 37,   # Every 20.000 hours
    "25000 HOURS": 41,   # Every 25.000 hours
    "40000 HOURS": 38,   # Every 40.000 hours
    "100000 HOURS":47,   # Every 100.000 hours
}

# Mappatura LivelloManMM → Maintenance_Level.id
LEVEL_MAP = {
    "1° LIVELLO": 1,   # C - On-board by crew operators
    "2° LIVELLO": 2,   # O - On-board by crew operators and technicians
    "3° LIVELLO": 3,   # F - On-board by base navy maintenance staff
    "4° LIVELLO": 4,   # H - On-board at naval base by Industry
    "5° LIVELLO": 5,   # D - Dockyard or Industry
    "6° LIVELLO": 6,   # L - Industry by industry qualified personnel
    # Alias alternativi che potresti trovare nei file Excel
    "LIVELLO C":  1,
    "LIVELLO O":  2,
    "LIVELLO F":  3,
    "LIVELLO H":  4,
    "LIVELLO D":  5,
    "LIVELLO L":  6,
    "C":          1,
    "O":          2,
    "F":          3,
    "H":          4,
    "D":          5,
    "L":          6,
}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def safe(val, default=None):
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except Exception:
        pass
    return val

def safe_str(val, max_len=None):
    v = safe(val)
    if v is None:
        return None
    s = str(v).strip()
    if max_len:
        s = s[:max_len]
    return s if s else None

def safe_int(val):
    v = safe(val)
    if v is None:
        return None
    try:
        return int(float(v))
    except Exception:
        return None

def safe_float(val):
    v = safe(val)
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None

def normalize_periodicity(val):
    import re
    if not val:
        return None
    s = str(val).strip().upper()
    s = s.replace(" MONTH", " MONTHS").replace("MONTHSS", "MONTHS")
    s = s.replace(" YEAR", " YEARS").replace("YEARSS", "YEARS")
    s = s.replace(" WEEK", " WEEKS").replace("WEEKSS", "WEEKS")
    s = s.replace(" DAY", " DAYS").replace("DAYSS", "DAYS")
    s = s.replace(" HOUR", " HOURS").replace("HOURSS", "HOURS")
    import re as _re
    s = _re.sub(r"(\d)\.(\d{3})", r"\1\2", s)
    return s.strip()

def parse_quantity(qty_str):
    """Estrae il numero da '1.0 EA', '2.0 EA' ecc."""
    if not qty_str:
        return None
    try:
        return float(str(qty_str).split()[0])
    except Exception:
        return None

# ─────────────────────────────────────────────
# CARICAMENTO EXCEL
# ─────────────────────────────────────────────
print(f"[{datetime.now().strftime('%H:%M:%S')}] Caricamento Excel: {EXCEL_FILE}")

# Leggiamo con openpyxl per leggere i colori di sfondo
from openpyxl import load_workbook
wb = load_workbook(EXCEL_FILE)
ws = wb.active

print("Primi 10 colori:")
for i in range(2, 12):
    cell = ws.cell(row=i, column=1)
    fg = cell.fill.fgColor
    print(f"  Row {i}: type={fg.type}, rgb={fg.rgb if fg.type=='rgb' else 'N/A'}")

COLOR_GREEN   = "FFC5E0B3"
COLOR_YELLOW  = "FFFFE598"
COLOR_BLUE    = "FFDEEAF6"
COLOR_GREEN2  = "FF92D050"   # verde scuro (pdm34)
COLOR_SALMON  = "FFFAE2D5"   # salmone/arancione (pdm34)

MAIN_COLORS = {COLOR_GREEN, COLOR_YELLOW, COLOR_BLUE, COLOR_GREEN2, COLOR_SALMON}

# Leggiamo anche con pandas per avere i dati strutturati
df = pd.read_excel(EXCEL_FILE, sheet_name="Sheet1")
print(f"[{datetime.now().strftime('%H:%M:%S')}] Righe totali: {len(df)}")

def get_row_color(excel_row_num):
    """Restituisce il colore RGB della colonna A per la riga Excel (1-based)."""
    cell = ws.cell(row=excel_row_num, column=1)
    fg = cell.fill.fgColor
    if fg.type == 'rgb':
        return fg.rgb
    return "00000000"

# ─────────────────────────────────────────────
# RAGGRUPPAMENTO: manutenzione + figli
# Riga verde  (FF92D050) + NrApp non-null → manutenzione
# Riga azzurra (FFDEEAF6)                 → SALTA (non è manutenzione)
# Riga trasparente (00000000)             → figlio della manutenzione corrente
# ─────────────────────────────────────────────
groups = []
current = None
skipped_blue = 0

for df_idx, row in df.iterrows():
    excel_row = df_idx + 2
    color = get_row_color(excel_row)

    if color in MAIN_COLORS and pd.notna(row['NrApp']):
        current = {'row': row, 'idx': df_idx, 'excel_row': excel_row, 'children': []}
        groups.append(current)
    elif current is not None and pd.isna(row['NrApp']):
        current['children'].append({'row': row, 'idx': df_idx})

print(f"[{datetime.now().strftime('%H:%M:%S')}] Manutenzioni trovate: {len(groups)} | Righe azzurre saltate: {skipped_blue}")

# ─────────────────────────────────────────────
# CONNESSIONE DB
# ─────────────────────────────────────────────
print(f"[{datetime.now().strftime('%H:%M:%S')}] Connessione al database...")
try:
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASSWORD,
        database=DB_NAME, charset="utf8mb4",
        autocommit=False
    )
    cursor = conn.cursor()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Connessione OK")
except Exception as e:
    print(f"ERRORE connessione: {e}")
    sys.exit(1)

# ─────────────────────────────────────────────
# CACHE: carica ElementModel LCN→ID già nel DB
# (per collegare la manutenzione all'apparato)
# ─────────────────────────────────────────────
print(f"[{datetime.now().strftime('%H:%M:%S')}] Caricamento mappa LCN→ElementModel dal DB...")
cursor.execute("SELECT id, LCN, ESWBS_code FROM ElementModel WHERE ship_model_id IN (SELECT ship_model_id FROM Ship WHERE id = %s)", (SHIP_ID,))
lcn_to_em_id = {}
for row in cursor.fetchall():
    em_id, lcn, eswbs = row
    if lcn:
        lcn_to_em_id[str(lcn).strip()] = em_id
    if eswbs:
        lcn_to_em_id[str(eswbs).strip()] = em_id
print(f"  → {len(lcn_to_em_id)} ElementModel caricati")

# Cache Consumables PN→ID (evita duplicati)
cursor.execute("SELECT ID, ConsumableArticleCode, Commercial_Name FROM Consumables")
consumable_cache = {}
for row in cursor.fetchall():
    cid, code, name = row
    if code:
        consumable_cache[str(code).strip()] = cid
    if name:
        consumable_cache[str(name).strip()] = cid

# Cache Tools (PN→ID)
cursor.execute("SELECT ID, Part_Number_OEM, Tool_name FROM Tools WHERE ship_id = %s", (SHIP_ID,))
tool_cache = {}
for row in cursor.fetchall():
    tid, pn, name = row
    if pn:
        tool_cache[str(pn).strip()] = tid
    if name:
        tool_cache[str(name).strip()] = tid

# Cache Spare (Serial→ID)
cursor.execute("SELECT ID, Serial_number, Part_name FROM Spare WHERE ship_id = %s", (SHIP_ID,))
spare_cache = {}
for row in cursor.fetchall():
    sid, serial, name = row
    if serial:
        spare_cache[str(serial).strip()] = sid
    if name:
        spare_cache[str(name).strip()] = sid

print(f"  → Cache: {len(consumable_cache)} consumabili, {len(tool_cache)} tools, {len(spare_cache)} spare")

stats = {
    "maint_ok": 0, "maint_skip": 0,
    "cons_new": 0, "cons_link": 0,
    "tool_new": 0, "tool_link": 0,
    "spare_new": 0, "spare_link": 0,
    "errors": []
}

# ─────────────────────────────────────────────
# PASS 1: INSERT Maintenance + collegamento figli
# ─────────────────────────────────────────────
print(f"\n[{datetime.now().strftime('%H:%M:%S')}] === Import Maintenance ===")

MAINT_SQL = """
    INSERT INTO Maintenance (
        id_ship, name,
        System_ElementModel_ID,
        End_Item_ElementModel_ID,
        Maintenance_Item_ElementModel_ID,
        Operational_Not_operational,
        Mean_elapsed_time_MELAP,
        Mean_Men_Hours_MMH,
        Personnel_no,
        RecurrencyType_ID, MaintenanceLevel_ID,
        Maintenance_type_id,
        Service_or_Maintenance_Manual_Link,
        `Service_or_Maintenance_manual_ParagraphPage`,
        Check_List,
        Maintenance_under_condition_description,
        Note
    ) VALUES (%s,%s, %s,%s,%s, %s, %s,%s,%s, %s,%s,%s, %s,%s, %s,%s,%s)
"""

def get_maintenance_type(norm_period):
    if not norm_period:
        return 1
    if "HOURS" in norm_period:
        return 2
    if "CYCLES" in norm_period:
        return 3
    if "ON FAULT" in norm_period:
        return 4
    if "ON CONDITION" in norm_period:
        return 5
    return 1
#print("Colonne Excel:", df.columns.tolist())

for group in groups:
    row = group['row']
    idx = group['idx']

    den_task    = safe_str(row['DenTask'], 255)
    eswbs_sys   = safe_str(row['ESWBS'], 20)
    eswbs_app   = safe_str(row['ESWBS Apparato'], 20)
    periodicita = safe_str(row['Periodicità'])
    livello     = safe_str(row['LivelloManMM']) or safe_str(row.get('LivelloManMM\n(NO)'))
    operativo   = safe_str(row['Operativo'], 100)
    nr_man      = safe_int(row['NR_Manutentori'])
    melap       = safe_str(row['MELAP (Minutes)'], 100)
    mmh         = safe_str(row['MMH (N item x MELAP x NR_Manutentori) - Minutes'], 100)
    doc_rif     = safe_str(row['DocumentoDiRiferimento'], 100)
    pagina      = safe_str(row['Pagina'], 100)
    applicab    = safe_str(row.get('Applicabilità'), 255)
    melap_orig  = safe_str(row.get('MelapOrigine (Minutes)'))
    period_orig = safe_str(row.get('Periodicità di origine'))

    note_parts = []
    if melap_orig:
        note_parts.append(f"MELAP origine: {melap_orig}")
    if period_orig:
        note_parts.append(f"Periodicità origine: {period_orig}")
    note = " | ".join(note_parts)[:100] if note_parts else None

    sys_em_id = lcn_to_em_id.get(eswbs_sys) if eswbs_sys else None
    app_em_id = lcn_to_em_id.get(eswbs_app) if eswbs_app else None

    norm_period   = normalize_periodicity(periodicita)
    recurrency_id = PERIODICITY_MAP.get(norm_period) if norm_period else None
    if periodicita and not recurrency_id:
        stats["errors"].append(f"Row {idx+2} WARNING: Periodicita '{periodicita}' non mappata (normalizzata: '{norm_period}')")

    level_id      = LEVEL_MAP.get(livello) if livello else None
    maint_type_id = get_maintenance_type(norm_period)
    check_list    = 1 if norm_period == "1 DAYS" else None

    try:
        cursor.execute(MAINT_SQL, (
            SHIP_ID, den_task,
            sys_em_id,
            app_em_id,
            app_em_id,
            operativo,
            melap,
            mmh,
            nr_man,
            recurrency_id, level_id,
            maint_type_id,
            doc_rif, pagina,
            check_list,
            applicab,
            note,
        ))
        maint_id = cursor.lastrowid
        stats["maint_ok"] += 1

        # ── Figli ──────────────────────────────────────────────────
        for child in group['children']:
            crow     = child['row']
            typology = safe_str(crow['TYPOLOGY'])
            pn       = safe_str(crow['PN'], 100)
            desc     = safe_str(crow['DESCRIPTION'], 255)
            cage     = safe_str(crow['CAGE'], 10)
            qty_raw  = safe_str(crow['QUANTITY'])
            qty      = parse_quantity(qty_raw)
            uom      = str(qty_raw).split()[-1] if qty_raw else 'EA'
            icc      = safe_str(crow['ICC'], 100)

            label = pn or desc or "N/D"

            if typology == "Consumable":
                # Cerca o crea Consumable
                cons_id = consumable_cache.get(label)
                if not cons_id:
                    cursor.execute("""
                        INSERT INTO Consumables (
                            Commercial_Name, ConsumableArticleCode,
                            ICC_Item_Category_Code, Consumable_quantity,
                            Consumable_unit_of_measure
                        ) VALUES (%s,%s,%s,%s,%s)
                    """, (desc or label, pn, icc, qty, uom))
                    cons_id = cursor.lastrowid
                    consumable_cache[label] = cons_id
                    stats["cons_new"] += 1

                cursor.execute("""
                    INSERT INTO Maintenance_ListConsumable
                        (Maintenance_List_ID, Consumable_ID,
                         Consumable_quantity, Consumable_quantity_Unit_of_measure)
                    VALUES (%s,%s,%s,%s)
                """, (maint_id, cons_id, qty, uom))
                stats["cons_link"] += 1

            elif typology == "Tool":
                # Cerca o crea Tool
                tool_id = tool_cache.get(label)
                if not tool_id:
                    cursor.execute("""
                        INSERT INTO Tools (
                            ship_id, Tool_name, Part_Number_OEM,
                            Original_description_OEM, ICC_Item_Category_Code
                        ) VALUES (%s,%s,%s,%s,%s)
                    """, (SHIP_ID, desc or label, pn, desc, icc))
                    tool_id = cursor.lastrowid
                    tool_cache[label] = tool_id
                    stats["tool_new"] += 1

                cursor.execute("""
                    INSERT INTO Maintenance_ListTools
                        (Maintenance_List_ID, `Tool ID`,
                         Tool_quantity, Tool_Quantity_Unit_of_measure)
                    VALUES (%s,%s,%s,%s)
                """, (maint_id, tool_id, qty, uom))
                stats["tool_link"] += 1

            elif typology == "Spare":
                # Cerca o crea Spare
                spare_id = spare_cache.get(label)
                if not spare_id:
                    serial = f"SPARE-MAINT-{pn or label}-{maint_id}"[:255]
                    cursor.execute("""
                        INSERT INTO Spare (
                            ship_id, Serial_number, Part_name, NSN
                        ) VALUES (%s,%s,%s,%s)
                    """, (SHIP_ID, serial, desc or label, pn))
                    spare_id = cursor.lastrowid
                    spare_cache[label] = spare_id
                    stats["spare_new"] += 1

                cursor.execute("""
                    INSERT INTO Maintenance_ListSpare
                        (Maintenance_List_ID, Spare_ID,
                         Spare_quantity, Spare_unit_of_measure)
                    VALUES (%s,%s,%s,%s)
                """, (maint_id, spare_id, safe_int(qty), uom))
                stats["spare_link"] += 1

    except Exception as e:
        conn.rollback()
        stats["maint_skip"] += 1
        stats["errors"].append(f"Row {idx+2} Maintenance '{den_task}': {e}")
        continue

    # ── JobExecution: crea istanza pianificata per questa manutenzione ──
    try:
        from datetime import timedelta
        import re as _re

        # ── Element: cerca per ESWBS Apparato (LCN completo) ──────────
        # lcn_to_em_id mappa LCN → ElementModel.id
        # Da ElementModel.id troviamo Element.id (ship_id + element_model_id)
        element_id = None
        if app_em_id:
            cursor.execute(
                "SELECT id FROM Element WHERE element_model_id = %s AND ship_id = %s LIMIT 1",
                (app_em_id, SHIP_ID)
            )
            el_row = cursor.fetchone()
            if el_row:
                element_id = el_row[0]
            else:
                # Fallback: prova con ESWBS sistema (3 cifre)
                sys_em_id_fb = lcn_to_em_id.get(eswbs_sys) if eswbs_sys else None
                if sys_em_id_fb:
                    cursor.execute(
                        "SELECT id FROM Element WHERE element_model_id = %s AND ship_id = %s LIMIT 1",
                        (sys_em_id_fb, SHIP_ID)
                    )
                    el_row2 = cursor.fetchone()
                    if el_row2:
                        element_id = el_row2[0]

        # ── ending_date: calcola dalla periodicità normalizzata ────────
        # Periodicità a calendario → somma giorni a oggi
        # Periodicità a ore/cicli  → NULL (non calcolabile a calendario)
        ending_date = None
        if norm_period:
            # Mappa diretta periodicità → giorni
            PERIOD_DAYS = {
                "1 DAYS": 1,       "1 WEEKS": 7,      "2 WEEKS": 14,
                "1 MONTHS": 30,    "2 MONTHS": 60,    "3 MONTHS": 90,
                "4 MONTHS": 120,   "6 MONTHS": 180,
                "1 YEARS": 365,    "2 YEARS": 730,    "2.5 YEARS": 912,
                "3 YEARS": 1095,   "4 YEARS": 1460,   "5 YEARS": 1825,
                "6 YEARS": 2190,   "8 YEARS": 2920,   "10 YEARS": 3650,
                "15 YEARS": 5475,
            }
            days = PERIOD_DAYS.get(norm_period)
            if days:
                ending_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
            elif "HOURS" not in norm_period and "CYCLES" not in norm_period:
                # Prova a estrarre numero + unità generica non mappata
                m = _re.match(r"(\d+(?:\.\d+)?)\s+(DAYS|WEEKS|MONTHS|YEARS)", norm_period)
                if m:
                    n, unit = float(m.group(1)), m.group(2)
                    mult = {"DAYS": 1, "WEEKS": 7, "MONTHS": 30, "YEARS": 365}
                    ending_date = (datetime.now() + timedelta(days=int(n * mult[unit]))).strftime("%Y-%m-%d %H:%M:%S")

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute("""
            INSERT INTO JobExecution (
                job_id, status_id, ship_id,
                element_eswbs_instance_id,
                recurrency_type_id,
                execution_date,
                starting_date,
                ending_date,
                execution_state
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            maint_id,       # job_id → Maintenance.id
            6,              # status_id = 6 (scheduled)
            SHIP_ID,
            element_id,     # Element dell'apparato (ESWBS Apparato → LCN → ElementModel → Element)
            recurrency_id,  # RecurrencyType
            now_str,        # execution_date = ora import
            now_str,        # starting_date = ora import
            ending_date,    # ending_date = starting + periodicità calendario (NULL per ore/cicli)
            "scheduled",    # execution_state
        ))
        stats["job_exec_ok"] = stats.get("job_exec_ok", 0) + 1
    except Exception as je:
        stats["errors"].append(f"Row {idx+2} JobExecution maint_id={maint_id}: {je}")
    # ────────────────────────────────────────────────────────────────────

    conn.commit()

# ─────────────────────────────────────────────
# RIEPILOGO
# ─────────────────────────────────────────────
cursor.close()
conn.close()

print(f"\n{'='*55}")
print(f"RIEPILOGO IMPORT MAINTENANCE — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'='*55}")
print(f"  Maintenance         : {stats['maint_ok']:>4} OK / {stats['maint_skip']} saltati")
print(f"  JobExecution creati : {stats.get('job_exec_ok', 0):>4}")
print(f"  Consumabili nuovi   : {stats['cons_new']:>4} | Collegati: {stats['cons_link']}")
print(f"  Tools nuovi         : {stats['tool_new']:>4} | Collegati: {stats['tool_link']}")
print(f"  Spare nuovi         : {stats['spare_new']:>4} | Collegati: {stats['spare_link']}")
print(f"  Errori              : {len(stats['errors'])}")

with open("import_maintenance_log.txt", "w", encoding="utf-8") as f:
    f.write(f"Import Maintenance — {datetime.now()}\n")
    f.write(f"Maintenance: {stats['maint_ok']} OK / {stats['maint_skip']} saltati\n")
    f.write(f"JobExecution creati: {stats.get('job_exec_ok', 0)}\n")
    f.write(f"Consumabili nuovi: {stats['cons_new']} | Collegati: {stats['cons_link']}\n")
    f.write(f"Tools nuovi: {stats['tool_new']} | Collegati: {stats['tool_link']}\n")
    f.write(f"Spare nuovi: {stats['spare_new']} | Collegati: {stats['spare_link']}\n\n")
    f.write("--- ERRORI ---\n")
    for e in stats["errors"]:
        f.write(e + "\n")

print(f"\nLog: import_maintenance_log.txt")
if stats["errors"]:
    print(f"\nErrori:")
    for e in stats["errors"]:
        print(f"  {e}")